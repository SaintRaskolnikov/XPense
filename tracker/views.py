from datetime import datetime, timedelta
from django.utils.timezone import now, make_aware, is_naive
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponseForbidden
from django.db import transaction as db_transaction
from django.contrib import messages
from django.core.exceptions import ValidationError
from .models import Transaction, Contribution, Subscription, Goals
from .forms import TransactionForm, ContributionForm, SubscriptionForm, GoalForm
from user.models import Team
from datetime import datetime, date


@login_required
def dashboard(request):
    today = now()

    # Default start and end dates
    default_start_date = today.replace(day=1)
    default_end_date = today

    # Parse start and end dates from GET params
    start_date_str = request.GET.get('start', default_start_date.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end', default_end_date.strftime('%Y-%m-%d'))

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        start_date = default_start_date
        start_date_str = default_start_date.strftime('%Y-%m-%d')

    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
    except (ValueError, TypeError):
        end_date = default_end_date
        end_date_str = default_end_date.strftime('%Y-%m-%d')

    # Fetch transactions and contributions
    transactions_pure = Transaction.objects.filter(user=request.user, team__isnull=True)
    contributions = Contribution.objects.filter(user=request.user, excluded=False)

    # Apply date filters
    if start_date:
        transactions_pure = transactions_pure.filter(date__gte=start_date)
        contributions = contributions.filter(transaction__date__gte=start_date)
    if end_date:
        transactions_pure = transactions_pure.filter(date__lte=end_date)
        contributions = contributions.filter(transaction__date__lte=end_date)

    # Prepare subscription warnings
    subs_renewal_warning = []
    subscriptions = Subscription.objects.filter(user=request.user, is_active=True)
    for subscription in subscriptions:
        next_due_date = subscription.get_next_transaction_date()
        if next_due_date:
            # If the next_due_date is a datetime.date object
            if isinstance(next_due_date, date):
                # Convert the date to datetime at midnight
                next_due_date = datetime.combine(next_due_date, datetime.min.time())
            
            # Now check if it's naive (it will be a datetime object by now)
            if is_naive(next_due_date):
                next_due_date = make_aware(next_due_date)  # Make it aware if naive
            
            # Compare with today's date
            if next_due_date <= today:
                subs_renewal_warning.append(subscription)
    
    # Calculate balance
    current_balance = calculate_balance(transactions_pure, contributions)

    # Combine and sort transactions and contributions
    transactions = merge_and_sort_transactions(transactions_pure, contributions)

    # Render the template
    return render(request, 'dashboard.html', {
        'current_balance': current_balance,
        'transactions': transactions,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'subs_renewal_warning': subs_renewal_warning,
    })


# Helper Function 1: Calculate Balance
def calculate_balance(transactions_pure, contributions):
    balance = 0
    for transaction in transactions_pure:
        if transaction.transaction_type == 'add':
            balance += transaction.amount
        elif transaction.transaction_type == 'expense':
            balance -= transaction.amount

    for contribution in contributions:
        if contribution.transaction.transaction_type == 'add':
            balance += contribution.amount
        elif contribution.transaction.transaction_type == 'expense':
            balance -= contribution.amount

    return balance


# Helper Function 2: Merge and Sort Transactions
def merge_and_sort_transactions(transactions_pure, contributions):
    combined_transactions = list(transactions_pure) + list(contributions)
    return sorted(
        combined_transactions,
        key=lambda x: x.date if isinstance(x, Transaction) else x.transaction.date,
        reverse=True
    )



def get_balance_data(request):
    # Get today's date
    today = now()

    # Get the date 30 days ago
    thirty_days_ago = today - timedelta(days=30)

    # Fetch transactions and contributions for the user within the last 30 days
    transactions_pure = Transaction.objects.filter(user=request.user, team__isnull=True, date__gte=thirty_days_ago).order_by('date')
    contributions = Contribution.objects.filter(user=request.user, transaction__date__gte=thirty_days_ago, excluded=False).order_by('transaction__date')

    # Initialize balance
    current_balance = 0
    balance_data = []

    # Iterate from the oldest date to the newest (30th day ago to today)
    for i in range(31):
        # For each day, get the date starting from the 30th day ago
        day = thirty_days_ago + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day.replace(hour=23, minute=59, second=59, microsecond=999999)

        # Filter transactions for that specific day
        daily_transactions = transactions_pure.filter(date__range=[day_start, day_end])
        daily_contributions = contributions.filter(transaction__date__range=[day_start, day_end])

        # Process daily transactions
        daily_balance = current_balance
        for transaction in daily_transactions:
            if transaction.transaction_type == 'add':
                daily_balance += transaction.amount
            elif transaction.transaction_type == 'expense':
                daily_balance -= transaction.amount

        for contribution in daily_contributions:
            if contribution.transaction.transaction_type == 'add':
                daily_balance += contribution.amount
            elif contribution.transaction.transaction_type == 'expense':
                daily_balance -= contribution.amount

        # Store the balance for the current day
        balance_data.append({
            'date': day.strftime('%Y-%m-%d'),
            'balance': daily_balance
        })

        # Set current balance to the calculated balance for this day
        current_balance = daily_balance

    return JsonResponse(balance_data, safe=False)




@login_required
def add_transaction(request):
    category_choices = Transaction.load_choices('category.json', language=request.user.language)  # Assuming you're loading categories from a JSON
    
    if request.method == 'POST':
        amount = request.POST.get('amount')
        amount = amount.replace(',', '.')  # Convert comma to period
        amount = float(amount)  # Convert to float after the replacement
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():

            with db_transaction.atomic():  # Ensure atomicity
                # Create the transaction object but don't save it yet
                transaction = form.save(commit=False)
                transaction.user = request.user  # Assign current user as creator
                transaction.save()  # Save transaction first
                transaction_hash = transaction.transaction_hash

                if transaction.team:
                    return redirect('tracker:add_contribution', transaction_hash=transaction_hash)  # Replace with your URL name for success page
                else:
                    # Update goals with the same category
                    goals_with_same_category = Goals.objects.filter(user=request.user, category=transaction.category)
                    for goal in goals_with_same_category:
                        
                        goal.current_amount += transaction.amount
                        if goal.current_amount >= goal.target_amount:
                            goal.is_completed = True
                        goal.save()

        return redirect('tracker:dashboard')

    else:
        form = TransactionForm(user=request.user)

    teams = Team.objects.filter(users=request.user)  # Assuming teams the user belongs to
    return render(request, 'add_transaction.html', {'form': form, 'teams': teams, 'category_choices': category_choices})


@login_required
def add_contribution(request, transaction_hash):
    transaction = Transaction.objects.get(transaction_hash=transaction_hash)
    members = transaction.team.users.all()  # Assuming team.users is a related manager
    members_count = members.count()
    amount_per_member = transaction.amount / members_count

    if request.method == 'POST':
        for member in members:
            # Get the amount for each member from the POST data
            amount = request.POST.get(f'amount_{member.id}')

            # Check if a contribution for this user and transaction already exists
            existing_contribution = Contribution.objects.filter(transaction=transaction, user=member, excluded=False).first()
            if existing_contribution:
                # Delete the existing contribution
                existing_contribution.delete()
                

            # Prepare the form data to create the contribution
            form_data = {
                'amount': amount,
                'transaction': transaction.id,
                'user': member.id,
            }

            # Create a new form with the provided data
            form = ContributionForm(form_data)

            # Check if the form is valid before saving
            if form.is_valid():
                
                contribution = form.save(commit=False)
                contribution.user = member
                contribution.transaction = transaction
                contribution.save()
                                # Update goals with the same category
                goals_with_same_category = Goals.objects.filter(user=member, category=transaction.category)
                for goal in goals_with_same_category:
                    
                    goal.current_amount += transaction.amount
                    if goal.current_amount >= goal.target_amount:
                        goal.is_completed = True
                    goal.save()
                    

        return redirect('tracker:dashboard')  # Redirect to a success page after saving

    else:
        form = ContributionForm()

    # Pass the form and necessary context data to the template
    return render(request, 'add_contribution.html', {
        'form': form,
        'members': members,
        'amount_per_member': amount_per_member,
        'transaction': transaction,
    })



@login_required
def get_team_members(request, pk):
    try:
        team = Team.objects.get(pk=pk)
        members = team.users.all()  # Assuming a ManyToMany relationship named 'users'
        members_data = [{"id": member.id, "username": member.username} for member in members]
        
        return JsonResponse({"members": members_data}, status=200)
    except Team.DoesNotExist:
        return JsonResponse({"error": "Team not found."}, status=404)

@login_required
def edit_transaction(request, transaction_hash):
    # Get the transaction object or return a 404 if not found
    transaction = get_object_or_404(Transaction, transaction_hash=transaction_hash)

    # Check if the current user is allowed to edit this transaction
    if transaction.user != request.user:
        return HttpResponseForbidden("You are not allowed to edit this transaction.")

    # If the form is submitted
    if request.method == "POST":
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.save()
            if transaction.team:
                    return redirect('tracker:add_contribution', transaction_hash=transaction_hash)  # Redirect to the dashboard or another page
    else:
        # Pre-fill the form with the transaction's current data
        form = TransactionForm(instance=transaction)

    teams = Team.objects.filter(users=request.user)  # Fetch teams the user is part of
    category_choices = Transaction.load_choices('category.json', language=request.user.language)  # Assuming you have category choices in the model

    return render(request, "edit_transaction.html", {
        "form": form,
        "transaction": transaction,
        "teams": teams,
        "category_choices": category_choices,
    })


@login_required
def delete_transaction(request, transaction_hash):
    # Fetch the transaction by transaction_hash
    transaction = get_object_or_404(Transaction, transaction_hash=transaction_hash, user=request.user)

    if request.method == 'POST':
        transaction.delete()
        messages.success(request, "Transaction deleted successfully!")
        return redirect('tracker:dashboard')  # Redirect to dashboard after delete

    return render(request, 'delete_transaction.html', {'transaction': transaction})

def exclude_contribution(request, pk):
    contribution = get_object_or_404(Contribution, pk=pk)
    
    if request.method == 'POST':
        # Exclude the contribution and redirect to the dashboard
        contribution.exclude()
        return redirect('tracker:dashboard')
    
    # Render a confirmation page for GET requests
    return render(request, 'exclude_contribution.html', {'contribution': contribution})

@login_required
def goals_progress(request):
    goals = Goals.objects.filter(user=request.user)

    return render(request, 'list_goals.html', context={'goals': goals})

@login_required
def create_goal(request):
    if request.method == "POST":
        form = GoalForm(request.POST, user=request.user)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user  # Assign the current logged-in user to the goal
            try:
                goal.save()  # This will trigger the validation in the model's save method
                messages.success(request, 'Goal created successfully!')
                return redirect('tracker:goals_progress')  # Redirect to a goal list or another appropriate page
            except ValidationError as e:
                # If a ValidationError is raised, add it to the form's errors
                form.add_error(None, e.message)
                messages.error(request, e.message)  # Pass the error message to the user
        else:
            messages.error(request, 'There was an error with your form submission. Please try again.')
    else:
        form = GoalForm(user=request.user)

    category_choices = Goals.load_choices('category.json', language=request.user.language) 
    return render(request, 'add_goal.html', {'form': form, "category_choices": category_choices})

@login_required
def edit_goal(request, pk):
    goal = get_object_or_404(Goals, id=pk, user=request.user)
    
    if request.method == "POST":
        form = GoalForm(request.POST, instance=goal, user=request.user)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Goal updated successfully!')
                return redirect('tracker:goals_progress')  # Redirect to a list or detail page
            except ValidationError as e:
                # If a ValidationError is raised, add it to the form's errors
                form.add_error(None, e.message)
                messages.error(request, e.message)  # Pass the error message to the user
        else:
            messages.error(request, 'There was an error with your form submission. Please try again.')
    else:
        form = GoalForm(instance=goal, user=request.user)

    category_choices = Goals.load_choices('category.json', language=request.user.language) 
    return render(request, 'edit_goal.html', {'form': form, 'goal': goal, "category_choices": category_choices})

def delete_goal(request, pk):
    goal = get_object_or_404(Goals, id=pk)

    if request.method == 'POST':
        goal.delete()
        messages.success(request, 'Goal deleted successfully.')
        return redirect('tracker:goals_progress')  # Redirect to the goal list after deletion

    return render(request, 'delete_goal.html', {'goal': goal})

@login_required
def get_graph_data(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    

    # Parse the start and end date if provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        
    else:
        # Default to the last 7 days if no date range is provided
        end_date = now()
        start_date = end_date - timedelta(days=7)

    # Get the selected team IDs from the request
    selected_team_ids = request.GET.getlist('team_ids')
    selected_team_ids = [int(team_id) for team_id in selected_team_ids]
    

    # Filter transactions and contributions based on the selected date range and teams
    transactions_pure = Transaction.objects.filter(
        user=request.user,
        team__id__in=selected_team_ids,
        date__range=[start_date, end_date]
    ).order_by('date')

   

    contributions = Contribution.objects.filter(
        user=request.user,
        excluded=False,
        transaction__date__range=[start_date, end_date],
        transaction__team__id__in=selected_team_ids
    ).order_by('transaction__date')


    # Initialize team totals (remove category handling)
    team_totals = {team.id: 0 for team in Team.objects.filter(id__in=selected_team_ids)}

    # Loop through transactions and contributions to accumulate the total spent per team
    for transaction in transactions_pure:
        if transaction.transaction_type == 'add':
            team_totals[transaction.team.id] += transaction.amount
        elif transaction.transaction_type == 'expense':
            team_totals[transaction.team.id] -= transaction.amount

    for contribution in contributions:
        if contribution.transaction.transaction_type == 'add':
            team_totals[contribution.transaction.team.id] += contribution.amount
        elif contribution.transaction.transaction_type == 'expense':
            team_totals[contribution.transaction.team.id] -= contribution.amount

    # Prepare the final data for the response
    total_spent_per_team = [
        {
            'team': Team.objects.get(id=team_id).name,  # Assuming you want the team name
            'total_spent': team_totals[team_id]
        }
        for team_id in selected_team_ids
    ]



    return JsonResponse({
        'total_spent_per_team': total_spent_per_team
    })

@login_required
def create_subscriptions(request):
    category_choices = Subscription.load_choices('category.json', request.user.language)    
    if request.method == 'POST':
        form = SubscriptionForm(request.POST, user=request.user)
        if form.is_valid():
            subscription = form.save(commit=False)
            subscription.user = request.user
            subscription.save()
            return redirect('tracker:user_subscriptions')
    else:
        form = SubscriptionForm(user=request.user)

    subscriptions = Subscription.objects.filter(user=request.user)
    return render(request, 'add_subscription.html', {'form': form, 'subscriptions': subscriptions, 'category_choices': category_choices})

@login_required
def edit_subscription(request, subscription_id):
    # Get the subscription object or raise a 404 if not found
    category_choices = Subscription.load_choices('category.json', request.user.language)
    subscription = get_object_or_404(Subscription, id=subscription_id)
    
    # Check if the subscription belongs to the current user
    if subscription.user != request.user:
        raise Http404("You are not authorized to edit this subscription.")

    # If the form is submitted
    if request.method == 'POST':
        form = SubscriptionForm(request.POST, instance=subscription, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('tracker:user_subscriptions')
    else:
        form = SubscriptionForm(instance=subscription, user=request.user)

    return render(request, 'edit_subscription.html', {
        'form': form,
        'subscription': subscription,
        'category_choices': category_choices,
    })

@login_required
def renew_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk)
    if subscription.user != request.user:
        raise Http404("You are not authorized to renew this subscription.")
    else:

        if request.method == 'POST':
            subscription.renew()
            
            Transaction.objects.create(
                user=subscription.user,
                category=subscription.category,
                amount=subscription.amount,
                description=subscription.name,
                transaction_type='expense',
                date=subscription.get_next_transaction_date(),
            )
            return redirect('tracker:dashboard')  # Redirect to the dashboard or another page


@login_required
def cancel_subscription(request, pk):
    subscription = get_object_or_404(Subscription, pk=pk)
    if subscription.user != request.user:
        raise Http404("You are not authorized to cancel this subscription.")
    else:
        if request.method == 'POST':
            subscription.cancel()
            return redirect('tracker:dashboard')


@login_required
def user_subscriptions(request):
    # Fetch all subscriptions that belong to the logged-in user
    subscriptions = Subscription.objects.filter(user=request.user)
    
    # Calculate monthly and annual expected spend
    daily_spend = 0
    weekly_spend = 0 + daily_spend
    monthly_spend = 0 + weekly_spend
    annual_spend = 0

    for subscription in subscriptions:
        if subscription.periodicity == 'daily':
            daily_spend += subscription.amount * 30
        elif subscription.periodicity == 'weekly':
            weekly_spend += subscription.amount * 4

        elif subscription.periodicity == 'monthly':
            monthly_spend += subscription.amount
        elif subscription.periodicity == 'annual':
            annual_spend += subscription.amount / 12  # Divide by 12 to get monthly equivalent

    # Calculate annual total spend (monthly spend + annual subscriptions)
    weekly_spend += daily_spend
    monthly_spend += weekly_spend
    total_annual_spend = monthly_spend * 12 + annual_spend

    return render(request, 'list_subscriptions.html', {
        'subscriptions': subscriptions,
        'daily_spend': daily_spend,
        'weekly_spend': weekly_spend,
        'monthly_spend': monthly_spend,
        'annual_spend': annual_spend,
        'total_annual_spend': total_annual_spend,
    })


@login_required
def delete_subscription(request, subscription_id):
    subscription = get_object_or_404(Subscription, id=subscription_id)
    
    if request.method == 'POST':
        subscription.delete()
        return redirect('tracker:user_subscriptions')  # Redirect to subscription list or desired page

    return render(request, 'delete_subscription.html', {'subscription': subscription})

@login_required
def delete_transaction(request, transaction_hash):
    transaction = get_object_or_404(Transaction, transaction_hash=transaction_hash)
    
    if request.method == 'POST':
        transaction.delete()
        return redirect('tracker:dashboard')  # Redirect to subscription list or desired page

    return render(request, 'delete_transaction.html', {'transaction': transaction})

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.timezone import localtime

@login_required
def export_transactions(request, interval):
    # Get the start and end dates based on the interval
    if interval == 'current_month':
        start_date = now().replace(day=1)
    elif interval == 'current_year':
        start_date = now().replace(month=1, day=1)
    elif interval == 'last_year':
        start_date = now().replace(year=now().year - 1, month=1, day=1)
    elif interval == 'last_month':
        if now().month == 1:
            start_date = now().replace(year=now().year - 1, month=12, day=1)
        else:
            start_date = now().replace(month=now().month - 1, day=1)
    else:
        start_date = None

    # Filter transactions and contributions based on the interval
    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        team__isnull=True
    ).order_by('-date')

    contributions = Contribution.objects.filter(
        user=request.user,
        excluded=False,
        transaction__date__gte=start_date
    ).order_by('transaction__date')

    full_list_transactions = list(transactions) + list(contributions)
    transactions = sorted(
        full_list_transactions,
        key=lambda x: x.date if isinstance(x, Transaction) else x.transaction.date,  # Check if it's a Transaction or Contribution and use the appropriate date
        reverse=True
    )
    # Create a workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Write headers
    headers = ['Description', 'Amount', 'Category', 'Date', 'Transaction Type']
    ws.append(headers)

    # Write data
    for item in transactions:
        # Ensure the date is naive (removing tzinfo if it's timezone-aware)
        date_value = item.date if isinstance(item, Transaction) else item.transaction.date
        if date_value.tzinfo:
            date_value = localtime(date_value).replace(tzinfo=None)  # Convert to naive datetime

        row = [
            item.description if isinstance(item, Transaction) else item.transaction.description,
            item.amount if isinstance(item, Transaction) else item.transaction.amount,
            item.category if isinstance(item, Transaction) else item.transaction.category,
            date_value,
            item.transaction_type if isinstance(item, Transaction) else item.transaction.transaction_type,
        ]
        ws.append(row)

    # Set column widths for better readability
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="transactions_{interval}.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response



@login_required
def graphs(request):
    current_year = datetime.now().year

    # Fetch transactions and contributions for the user with the date filter
    transactions_pure = Transaction.objects.filter(user=request.user, team__isnull=True, date__year=current_year)
    contributions = Contribution.objects.filter(user=request.user, transaction__date__year=current_year, excluded=False)

    # Calculate totals
    total_income = sum(transaction.amount for transaction in transactions_pure if transaction.transaction_type == 'add')
    total_expense = sum(transaction.amount for transaction in transactions_pure if transaction.transaction_type == 'expense')
    total_income += sum(contribution.amount for contribution in contributions if contribution.transaction.transaction_type == 'add')
    total_expense += sum(contribution.amount for contribution in contributions if contribution.transaction.transaction_type == 'expense')

    profit = total_income - total_expense
    profit_ratio = (profit / total_income) * 100 if total_income > 0 else 0

    return render(request, 'graphs.html', {
        'total_income': total_income,
        'total_expense': total_expense,
        'profit': profit,
        'profit_ratio': profit_ratio
    })


@login_required
def get_monthly_balance_data(request):
    current_year = datetime.now().year
    # Fetch transactions and contributions
    transactions_pure = Transaction.objects.filter(user=request.user, team__isnull=True, date__year=current_year)
    contributions = Contribution.objects.filter(user=request.user, transaction__date__year=current_year, excluded=False)

    monthly_balance = {month: {'income': 0, 'expense': 0} for month in range(1, 13)}

    for transaction in transactions_pure:
        month = transaction.date.month
        if transaction.transaction_type == 'add':
            monthly_balance[month]['income'] += transaction.amount
        elif transaction.transaction_type == 'expense':
            monthly_balance[month]['expense'] += transaction.amount

    for contribution in contributions:
        month = contribution.transaction.date.month
        if contribution.transaction.transaction_type == 'add':
            monthly_balance[month]['income'] += contribution.amount
        elif contribution.transaction.transaction_type == 'expense':
            monthly_balance[month]['expense'] += contribution.amount

    balance_data = [
        {
            'month': datetime(current_year, month, 1).strftime('%b'),
            'income': monthly_balance[month]['income'],
            'expense': monthly_balance[month]['expense'],
        }
        for month in range(1, 13)
    ]

    return JsonResponse(balance_data, safe=False)


@login_required
def get_expense_per_category(request):

    current_year = datetime.now().year
    

    transactions_pure = Transaction.objects.filter(user=request.user, team__isnull=True, date__year=current_year, transaction_type='expense')
    contributions = Contribution.objects.filter(user=request.user, transaction__date__year=current_year, transaction__transaction_type='expense', excluded=False)

    category_expenses = {}

    for transaction in transactions_pure:
        category = transaction.category
        category_expenses[category] = category_expenses.get(category, 0) + transaction.amount

    for contribution in contributions:
        category = contribution.transaction.category
        category_expenses[category] = category_expenses.get(category, 0) + contribution.amount

    expense_data = [
        {
            'category': category,
            'expense': float(amount),
        }
        for category, amount in category_expenses.items()
    ]

    return JsonResponse(expense_data, safe=False)